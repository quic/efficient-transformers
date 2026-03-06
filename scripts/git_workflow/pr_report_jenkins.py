#!/usr/bin/env python3
"""
PR Dashboard Report Generator for Jenkins

Generates comprehensive PR dashboard reports with:
- HTML email body with embedded table and pie chart
- Excel file with all PR data
- Standalone HTML page

Usage:
    export GITHUB_TOKEN=your_token
    export GITHUB_REPOSITORY=owner/repo
    python pr_report_jenkins.py

Output:
    - pr_dashboard.html (email body)
    - pr_dashboard.xlsx (Excel attachment)
    - pr_dashboard_full.html (standalone page)
"""

import json
import math
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Tuple, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Constants
API_BASE = "https://api.github.com"
API_ACCEPT = "application/vnd.github+json"
API_VERSION = "2022-11-28"
MAX_RETRIES = 3
ITEMS_PER_PAGE = 100

# Color palette for pie chart
CHART_COLORS = [
    "#4a90d9", "#e74c3c", "#2ecc71", "#f39c12", "#9b59b6",
    "#1abc9c", "#e67e22", "#3498db", "#e91e63", "#00bcd4",
    "#ff5722", "#607d8b", "#795548", "#9c27b0", "#4caf50",
]


# ── GitHub API Client ─────────────────────────────────────────────────────────


class GitHubAPIClient:
    """GitHub API client with retry logic and pagination support."""

    def __init__(self, token: str):
        self.token = token
        self.headers = {
            "Accept": API_ACCEPT,
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": API_VERSION,
        }

    def request(self, path: str, params: Dict = None) -> Tuple[Any, Dict]:
        """Make a GitHub API request with retry logic."""
        url = f"{API_BASE}{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params)

        req = urllib.request.Request(url, headers=self.headers)

        for attempt in range(MAX_RETRIES):
            try:
                with urllib.request.urlopen(req) as resp:
                    return json.loads(resp.read().decode("utf-8")), dict(resp.headers)
            except urllib.error.HTTPError as e:
                body = e.read().decode("utf-8", errors="replace")
                if e.code in (429, 403) and "rate limit" in body.lower():
                    wait = 60 * (attempt + 1)
                    print(f"Rate limited, waiting {wait}s...", file=sys.stderr)
                    time.sleep(wait)
                    continue
                print(f"HTTP {e.code}: {body[:300]}", file=sys.stderr)
                raise
            except urllib.error.URLError as e:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(5 * (attempt + 1))
                    continue
                raise

        raise RuntimeError(f"API request failed after {MAX_RETRIES} retries: {path}")

    def paginate(self, path: str, params: Dict = None) -> List[Any]:
        """Fetch all pages from a GitHub list endpoint."""
        page = 1
        results = []
        
        while True:
            p = dict(params or {})
            p.update({"per_page": ITEMS_PER_PAGE, "page": page})
            chunk, headers = self.request(path, p)
            
            if not chunk:
                break
                
            results.extend(chunk)
            
            if 'rel="next"' not in headers.get("Link", ""):
                break
                
            page += 1
            
        return results

    def paginate_check_runs(self, path: str, params: Dict = None) -> List[Any]:
        """Paginate check-runs endpoint (wrapped response format)."""
        page = 1
        results = []
        
        while True:
            p = dict(params or {})
            p.update({"per_page": ITEMS_PER_PAGE, "page": page})
            resp, headers = self.request(path, p)
            chunk = resp.get("check_runs", [])
            results.extend(chunk)
            
            if 'rel="next"' not in headers.get("Link", ""):
                break
                
            page += 1
            
        return results


# ── Data Models ───────────────────────────────────────────────────────────────


class PRData:
    """Data model for a Pull Request."""

    def __init__(self, pr_dict: Dict, client: GitHubAPIClient, owner: str, repo: str):
        self.number = pr_dict["number"]
        self.title = pr_dict.get("title", "")
        self.url = pr_dict.get("html_url", "")
        self.author = (pr_dict.get("user") or {}).get("login", "unknown")
        self.draft = "Yes" if pr_dict.get("draft") else "No"
        self.created_at = self._parse_iso(pr_dict["created_at"])
        self.age_days = (datetime.now(timezone.utc) - self.created_at).days
        self.head_sha = (pr_dict.get("head") or {}).get("sha")
        
        # Extract data from PR payload
        self.assignees = self._extract_assignees(pr_dict)
        self.labels = self._extract_labels(pr_dict)
        
        # Fetch additional data via API
        self.reviewers = self._fetch_reviewers(client, owner, repo)
        self.review_summary, self.pending_with = self._fetch_reviews(
            client, owner, repo, pr_dict
        )
        self.ci_status = self._fetch_ci_status(client, owner, repo)

    @staticmethod
    def _parse_iso(dt: str) -> datetime:
        return datetime.fromisoformat(dt.replace("Z", "+00:00"))

    @staticmethod
    def _is_bot(username: str) -> bool:
        return "[bot]" in username

    def _extract_assignees(self, pr_dict: Dict) -> str:
        assignees = [
            u["login"]
            for u in pr_dict.get("assignees") or []
            if not self._is_bot(u["login"])
        ]
        return ", ".join(assignees) if assignees else "—"

    def _extract_labels(self, pr_dict: Dict) -> str:
        labels = [lbl["name"] for lbl in pr_dict.get("labels") or []]
        return ", ".join(labels) if labels else "—"

    def _fetch_reviewers(self, client: GitHubAPIClient, owner: str, repo: str) -> str:
        rr, _ = client.request(f"/repos/{owner}/{repo}/pulls/{self.number}/requested_reviewers")
        users = [u["login"] for u in rr.get("users", []) if not self._is_bot(u["login"])]
        teams = [t["name"] for t in rr.get("teams", [])]
        reviewers = users + [f"team:{t}" for t in teams]
        return ", ".join(reviewers) if reviewers else "—"

    def _fetch_reviews(
        self, client: GitHubAPIClient, owner: str, repo: str, pr_dict: Dict
    ) -> Tuple[str, str]:
        reviews = client.paginate(f"/repos/{owner}/{repo}/pulls/{self.number}/reviews")
        
        # Summarize reviews
        latest = {}
        for r in sorted(reviews, key=lambda x: x.get("submitted_at") or ""):
            user = (r.get("user") or {}).get("login", "unknown")
            if self._is_bot(user):
                continue
            latest[user] = r.get("state", "UNKNOWN")

        approvers = sorted([u for u, s in latest.items() if s == "APPROVED"])
        changes_requested = sorted([u for u, s in latest.items() if s == "CHANGES_REQUESTED"])
        commenters = sorted([u for u, s in latest.items() if s == "COMMENTED"])
        dismissed = sorted([u for u, s in latest.items() if s == "DISMISSED"])

        # Build review summary
        parts = []
        if changes_requested:
            parts.append("Changes Requested: " + ", ".join(changes_requested))
        if approvers:
            parts.append("Approved: " + ", ".join(approvers))
        if commenters:
            parts.append("Commented: " + ", ".join(commenters))
        if dismissed:
            parts.append("Dismissed: " + ", ".join(dismissed))
        if not parts:
            parts.append("No reviews yet")
        
        review_summary = " / ".join(parts)
        
        # Determine pending with
        requested_reviewers = self.reviewers.split(", ") if self.reviewers != "—" else []
        pending_with = self._determine_pending_with(
            pr_dict, reviews, approvers, changes_requested, requested_reviewers
        )
        
        return review_summary, pending_with

    def _determine_pending_with(
        self,
        pr_dict: Dict,
        reviews: List,
        approvers: List,
        changes_requested: List,
        requested_reviewers: List,
    ) -> str:
        if pr_dict.get("draft"):
            return self.author

        if not changes_requested and not approvers:
            return ", ".join(requested_reviewers) if requested_reviewers else self.author

        if changes_requested:
            # Check if changes have been addressed
            last_cr_commits = {}
            for r in sorted(reviews, key=lambda x: x.get("submitted_at") or ""):
                user = (r.get("user") or {}).get("login", "unknown")
                if self._is_bot(user):
                    continue
                if r.get("state") == "CHANGES_REQUESTED":
                    last_cr_commits[user] = r.get("commit_id", "")

            unresolved = [
                r for r in changes_requested
                if last_cr_commits.get(r) == self.head_sha or not last_cr_commits.get(r)
            ]
            
            if unresolved:
                return self.author
            else:
                return ", ".join(changes_requested)

        if approvers and not changes_requested:
            return self.author

        return ", ".join(requested_reviewers) if requested_reviewers else self.author

    def _fetch_ci_status(self, client: GitHubAPIClient, owner: str, repo: str) -> str:
        if not self.head_sha:
            return "UNKNOWN"

        check_runs = client.paginate_check_runs(
            f"/repos/{owner}/{repo}/commits/{self.head_sha}/check-runs",
            params={"filter": "latest"},
        )

        if not check_runs:
            return "NONE"

        results = []
        for cr in sorted(check_runs, key=lambda x: x.get("name", "")):
            name = cr.get("name", "unknown")
            status = cr.get("status")
            conclusion = cr.get("conclusion")

            if status != "completed" or conclusion is None:
                state = "PENDING"
            elif conclusion in ("failure", "cancelled", "timed_out", "action_required", "stale"):
                state = "FAIL"
            elif conclusion in ("success", "neutral", "skipped"):
                state = "PASS"
            else:
                state = conclusion.upper()

            results.append(f"{name}: {state}")

        return " / ".join(results)

    def to_dict(self) -> Dict:
        """Convert to dictionary for Excel/HTML generation."""
        return {
            "number": self.number,
            "title": self.title,
            "url": self.url,
            "author": self.author,
            "assignee": self.assignees,
            "age_days": self.age_days,
            "draft": self.draft,
            "labels": self.labels,
            "reviewers": self.reviewers,
            "pending_with": self.pending_with,
            "review_summary": self.review_summary,
            "ci_status": self.ci_status,
        }


# ── Report Generators ─────────────────────────────────────────────────────────


class PieChartGenerator:
    """Generate SVG pie charts for PR distribution."""

    @staticmethod
    def generate(author_counts: Dict[str, int]) -> str:
        if not author_counts:
            return ""

        items = sorted(author_counts.items(), key=lambda x: -x[1])
        total = sum(v for _, v in items)

        cx, cy, r = 190, 190, 160
        legend_x = cx * 2 + 30
        row_h = 22
        svg_w = legend_x + 260
        svg_h = max(cy * 2, len(items) * row_h + 50)

        paths = []
        legends = []
        start_angle = -math.pi / 2

        for i, (author, count) in enumerate(items):
            angle = 2 * math.pi * count / total
            end_angle = start_angle + angle

            x1 = cx + r * math.cos(start_angle)
            y1 = cy + r * math.sin(start_angle)
            x2 = cx + r * math.cos(end_angle)
            y2 = cy + r * math.sin(end_angle)

            large_arc = 1 if angle > math.pi else 0
            color = CHART_COLORS[i % len(CHART_COLORS)]
            pct = count / total * 100

            path = (
                f'<path d="M {cx},{cy} L {x1:.2f},{y1:.2f} '
                f'A {r},{r} 0 {large_arc},1 {x2:.2f},{y2:.2f} Z" '
                f'fill="{color}" stroke="white" stroke-width="2">'
                f'<title>{author}: {count} PR{"s" if count != 1 else ""} ({pct:.1f}%)</title>'
                f'</path>'
            )
            paths.append(path)

            ly = 40 + i * row_h
            legend = (
                f'<rect x="{legend_x}" y="{ly}" width="14" height="14" '
                f'fill="{color}" rx="2"/>'
                f'<text x="{legend_x + 20}" y="{ly + 11}" '
                f'font-size="12" font-family="Arial, sans-serif" fill="#333">'
                f'{author}: {count} PR{"s" if count != 1 else ""} ({pct:.1f}%)'
                f'</text>'
            )
            legends.append(legend)

            start_angle = end_angle

        return f'''<svg width="{svg_w}" height="{svg_h}" xmlns="http://www.w3.org/2000/svg" style="font-family:Arial,sans-serif;">
  <text x="{cx}" y="20" text-anchor="middle" font-size="14" font-weight="bold" fill="#1a1a2e">
    PR Distribution by Author (Total: {total})
  </text>
  {''.join(paths)}
  <text x="{legend_x}" y="22" font-size="13" font-weight="bold" fill="#1a1a2e">Author</text>
  {''.join(legends)}
</svg>'''


class ExcelGenerator:
    """Generate Excel reports."""

    @staticmethod
    def generate(pr_data: List[Dict], output_file: str = "pr_dashboard.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PR Dashboard"

        # Styling
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "PR #", "Title", "URL", "Author", "Assignee", "Age (days)",
            "Draft", "Labels", "Reviewers", "Pending With", "Review Summary", "CI Status"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Data rows
        for row_num, pr in enumerate(pr_data, 2):
            for col_num, key in enumerate([
                "number", "title", "url", "author", "assignee", "age_days",
                "draft", "labels", "reviewers", "pending_with", "review_summary", "ci_status"
            ], 1):
                cell = ws.cell(row=row_num, column=col_num, value=pr[key])
                cell.border = border

        # Column widths
        widths = [8, 40, 50, 15, 15, 12, 8, 15, 15, 15, 30, 30]
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        wb.save(output_file)
        print(f"✓ Excel file: {output_file}", file=sys.stderr)


class HTMLGenerator:
    """Generate HTML reports."""

    @staticmethod
    def generate(
        pr_data: List[Dict],
        author_counts: Dict[str, int],
        repo_name: str,
        date_str: str,
    ) -> str:
        table_rows = "".join([
            f'''<tr>
            <td><a href="{pr['url']}">#{pr['number']}</a></td>
            <td>{pr['title']}</td>
            <td>{pr['author']}</td>
            <td>{pr['assignee']}</td>
            <td style="text-align:center">{pr['age_days']}</td>
            <td style="text-align:center">{pr['draft']}</td>
            <td>{pr['labels']}</td>
            <td>{pr['reviewers']}</td>
            <td>{pr['pending_with']}</td>
            <td>{pr['review_summary']}</td>
            <td>{pr['ci_status']}</td>
        </tr>'''
            for pr in pr_data
        ])

        pie_chart = PieChartGenerator.generate(author_counts)

        return f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: Arial, sans-serif; font-size: 13px; color: #333; max-width: 1400px; margin: 0 auto; padding: 16px; }}
        h1 {{ color: #1a1a2e; border-bottom: 2px solid #4a90d9; padding-bottom: 6px; font-size: 20px; }}
        .summary {{ background: #f8f9fa; padding: 12px; border-radius: 5px; margin: 16px 0; }}
        .summary table {{ border-collapse: collapse; width: 100%; }}
        .summary td {{ padding: 6px; }}
        .summary td:first-child {{ font-weight: bold; width: 150px; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 16px; }}
        th {{ background: #2c3e50; color: #fff; padding: 10px; text-align: left; font-size: 12px; }}
        td {{ border: 1px solid #ddd; padding: 8px; font-size: 12px; }}
        tr:nth-child(even) {{ background: #f8f9fa; }}
        tr:hover {{ background: #eaf2ff; }}
        a {{ color: #1a6bbf; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        .chart-container {{ margin: 24px 0; text-align: center; }}
    </style>
</head>
<body>
    <h1>📊 Open PR Dashboard — {repo_name}</h1>
    
    <div class="summary">
        <table>
            <tr><td>Report Date:</td><td>{date_str}</td></tr>
            <tr><td>Total Open PRs:</td><td><strong>{len(pr_data)}</strong></td></tr>
        </table>
    </div>

    <div class="chart-container">{pie_chart}</div>

    <table>
        <thead>
            <tr>
                <th>PR #</th><th>Title</th><th>Author</th><th>Assignee</th>
                <th>Age (days)</th><th>Draft</th><th>Labels</th><th>Reviewers</th>
                <th>Pending With</th><th>Review Summary</th><th>CI Status</th>
            </tr>
        </thead>
        <tbody>{table_rows}</tbody>
    </table>
</body>
</html>'''


# ── Main ──────────────────────────────────────────────────────────────────────


def main():
    # Validate environment
    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        print("ERROR: GITHUB_TOKEN environment variable not set", file=sys.stderr)
        sys.exit(1)

    repo_full = os.environ.get("GITHUB_REPOSITORY")
    if not repo_full or "/" not in repo_full:
        print("ERROR: GITHUB_REPOSITORY not set (format: owner/repo)", file=sys.stderr)
        sys.exit(1)

    owner, repo = repo_full.split("/", 1)
    date_str = datetime.now(timezone.utc).strftime("%B %d, %Y %H:%M UTC")

    print(f"Fetching PRs for {owner}/{repo}...", file=sys.stderr)

    # Initialize API client
    client = GitHubAPIClient(token)

    # Fetch PRs
    pulls = client.paginate(f"/repos/{owner}/{repo}/pulls", params={"state": "open"})
    print(f"Found {len(pulls)} open PRs", file=sys.stderr)

    # Process PRs
    pr_data_list = []
    author_counts = {}

    for pr_dict in pulls:
        pr = PRData(pr_dict, client, owner, repo)
        pr_data_list.append(pr.to_dict())
        
        if not PRData._is_bot(pr.author):
            author_counts[pr.author] = author_counts.get(pr.author, 0) + 1

    # Generate outputs
    print("Generating outputs...", file=sys.stderr)

    ExcelGenerator.generate(pr_data_list, "pr_dashboard.xlsx")
    
    html_content = HTMLGenerator.generate(
        pr_data_list, author_counts, f"{owner}/{repo}", date_str
    )
    
    Path("pr_dashboard.html").write_text(html_content)
    print("✓ HTML email body: pr_dashboard.html", file=sys.stderr)
    
    Path("pr_dashboard_full.html").write_text(html_content)
    print("✓ Standalone HTML: pr_dashboard_full.html", file=sys.stderr)

    print("\n✅ All outputs generated successfully!", file=sys.stderr)


if __name__ == "__main__":
    main()
